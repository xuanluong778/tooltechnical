"""
Keyword Research + Clustering APIs and HTML tool pages.
"""

from __future__ import annotations

import json
import os
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, Request, UploadFile
from fastapi.responses import HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel, Field
from sqlalchemy import desc
from sqlalchemy.orm import Session

from app.db import get_db
from app.models.keyword_cluster_project import KeywordClusterProject
from app.models.keyword_research_project import KeywordResearchProject
from app.models.user import User
from app.services.auth import get_billing_user, get_current_user
from app.services.user_scope import assert_job_access
from app.services.credits import (
    consume_credits,
    cost_cluster_async,
    cost_cluster_save,
    cost_cluster_sync,
    cost_import_excel,
    cost_research_run,
    volume_batch_charge,
)
from app.services.keyword_cluster_pipeline import build_keyword_cluster_api_response
from app.services.keyword_export import export_cluster_flat_to_excel, export_to_excel, export_to_google_sheets
from app.services.keyword_research_pipeline import build_keyword_research_api_response
from app.services.search_volume import enrich_keyword_volumes_cached

router = APIRouter(prefix="/keywords", tags=["keywords"])
templates = Jinja2Templates(directory="templates")


def _country_label(code: str) -> str:
    c = (code or "").lower()
    return {
        "vn": "Vietnam",
        "us": "United States",
        "gb": "United Kingdom",
    }.get(c, (code or "").upper() if len(c) == 2 else (code or ""))


class ResearchRunBody(BaseModel):
    seeds: list[str] = Field(..., min_length=1, max_length=50)
    domain: str | None = Field(default=None, max_length=200)
    url: str | None = Field(default=None, max_length=2000)
    engine: str = Field(default="google", max_length=32)
    language: str = Field(default="vi", max_length=16)
    country: str = Field(default="vn", max_length=16)
    device: str = Field(default="desktop", max_length=16)
    cluster_mode: str | None = Field(default=None, max_length=16, description="hybrid | tfidf | both")
    cluster_strictness: str | None = Field(default=None, max_length=16, description="strict | normal | loose")
    cluster_fetch_serp: bool | None = Field(default=None, description="Hybrid SERP fetch; None = env default")
    cluster_max_keywords: int | None = Field(default=None, description="Giới hạn KW cho clustering; server sẽ clamp an toàn")


class ClusterSaveBody(BaseModel):
    keywords: list[str] = Field(..., min_length=1, max_length=5000)
    fetch_serp: bool = Field(default=True)
    country: str = Field(default="vn", max_length=8)
    language: str = Field(default="vi", max_length=8)
    device: str = Field(default="desktop", max_length=16)
    cluster_strictness: str = Field(default="normal", max_length=16)
    url: str | None = Field(default=None, max_length=2000)
    result: dict = Field(..., description="Raw clustering response payload")


class VolumeBatchBody(BaseModel):
    keywords: list[str] = Field(..., min_length=1, max_length=5000)
    country: str = Field(default="vn", max_length=8)
    language: str = Field(default="vi", max_length=8)


@router.post("/volume/batch")
def api_keyword_volume_batch(
    body: VolumeBatchBody,
    db: Session = Depends(get_db),
    billing_user: User | None = Depends(get_billing_user),
) -> JSONResponse:
    kws = [str(s).strip() for s in body.keywords if str(s).strip()]
    if not kws:
        raise HTTPException(status_code=400, detail="At least one keyword required")
    rows = enrich_keyword_volumes_cached(kws, country=body.country, language=body.language)
    out: dict[str, dict] = {}
    for r in rows or []:
        kw = str(r.get("keyword") or "").strip()
        if not kw:
            continue
        out[kw] = {
            "search_volume": int(r.get("search_volume") or 0),
            "volume_source": str(r.get("volume_source") or ""),
            "confidence": float(r.get("confidence") or 0.0),
        }
    if billing_user:
        amt = volume_batch_charge(len(kws))
        if amt > 0:
            consume_credits(
                db,
                user_id=billing_user.id,
                amount=amt,
                reason="volume_batch",
                note=f"keywords={len(kws)}",
            )
            db.commit()
    return JSONResponse(content={"items": out})


def _cluster_label(keywords: list[str]) -> str:
    xs = [str(s).strip() for s in (keywords or []) if str(s).strip()]
    if not xs:
        return "Keyword clustering"
    head = " + ".join(xs[:8])
    if len(xs) > 8:
        head += " + …"
    return head[:2000]


@router.post("/cluster/save")
def api_cluster_save(
    body: ClusterSaveBody,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    billing_user: User | None = Depends(get_billing_user),
) -> JSONResponse:
    kws = [str(s).strip() for s in body.keywords if str(s).strip()]
    if not kws:
        raise HTTPException(status_code=400, detail="At least one keyword required")
    strict = (body.cluster_strictness or "normal").strip().lower()
    if strict not in ("strict", "normal", "loose"):
        strict = "normal"
    payload = body.result if isinstance(body.result, dict) else {}
    row = KeywordClusterProject(
        user_id=current_user.id,
        keywords_label=_cluster_label(kws),
        keywords_json=json.dumps(kws, ensure_ascii=False),
        language=(body.language or "vi").strip()[:16],
        country=(body.country or "vn").strip()[:16],
        device=(body.device or "desktop").strip()[:16],
        cluster_strictness=strict[:16],
        fetch_serp=1 if bool(body.fetch_serp) else 0,
        brand_url=(body.url or "").strip()[:2000],
        result_json=json.dumps(payload, ensure_ascii=False),
    )
    db.add(row)
    if billing_user:
        c = cost_cluster_save()
        if c > 0:
            consume_credits(db, user_id=billing_user.id, amount=c, reason="cluster_save", note=f"kw={len(kws)}")
    db.commit()
    db.refresh(row)
    return JSONResponse(content={"ok": True, "project_id": row.id})


@router.get("/cluster/history")
def api_cluster_history(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    limit: int = Query(50, ge=1, le=200),
) -> JSONResponse:
    rows = (
        db.query(KeywordClusterProject)
        .filter(KeywordClusterProject.user_id == current_user.id)
        .order_by(desc(KeywordClusterProject.created_at))
        .limit(limit)
        .all()
    )
    out: list[dict] = []
    for r in rows:
        out.append(
            {
                "id": r.id,
                "keywords_label": r.keywords_label,
                "language": r.language,
                "country": r.country,
                "device": r.device,
                "cluster_strictness": r.cluster_strictness,
                "fetch_serp": bool(r.fetch_serp),
                "brand_url": r.brand_url,
                "created_at": r.created_at.isoformat() if r.created_at else "",
            }
        )
    return JSONResponse(content={"projects": out})


@router.get("/cluster/history/{project_id}")
def api_cluster_history_item(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> JSONResponse:
    r = db.get(KeywordClusterProject, project_id)
    if not r or r.user_id != current_user.id or not r.result_json:
        raise HTTPException(status_code=404, detail="Project not found")
    return JSONResponse(
        content={
            "project": {
                "id": r.id,
                "keywords_label": r.keywords_label,
                "language": r.language,
                "country": r.country,
                "device": r.device,
                "cluster_strictness": r.cluster_strictness,
                "fetch_serp": bool(r.fetch_serp),
                "brand_url": r.brand_url,
                "created_at": r.created_at.isoformat() if r.created_at else "",
            },
            "result": json.loads(r.result_json),
            "keywords": json.loads(r.keywords_json) if r.keywords_json else [],
        }
    )


@router.delete("/cluster/history/{project_id}")
def api_cluster_history_delete(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> JSONResponse:
    r = db.get(KeywordClusterProject, project_id)
    if not r or r.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(r)
    db.commit()
    return JSONResponse(content={"ok": True})


@router.get("/research")
def api_keyword_research(
    seed_keyword: str = Query(..., min_length=1, max_length=2000),
    domain: str | None = Query(default=None, max_length=200),
    url: str | None = Query(default=None, max_length=2000),
    engine: str = Query("google", max_length=32),
    language: str = Query("vi", max_length=16),
    country: str = Query("vn", max_length=8),
    device: str = Query("desktop", max_length=16),
    cluster_mode: str | None = Query(default=None, max_length=16),
    cluster_strictness: str | None = Query(default=None, max_length=16),
    cluster_fetch_serp: str | None = Query(default=None, max_length=8),
    cluster_max_keywords: int | None = Query(default=None),
    db: Session = Depends(get_db),
    billing_user: User | None = Depends(get_billing_user),
) -> JSONResponse:
    cfs: bool | None = None
    if cluster_fetch_serp is not None and str(cluster_fetch_serp).strip() != "":
        cfs = str(cluster_fetch_serp).lower() in ("1", "true", "yes", "on")
    payload = build_keyword_research_api_response(
        seed_keyword=seed_keyword,
        domain=domain,
        url=url,
        engine=engine,
        language=language,
        country=country,
        device=device,
        cluster_mode=cluster_mode,
        cluster_strictness=cluster_strictness,
        cluster_fetch_serp=cfs,
        cluster_max_keywords=cluster_max_keywords,
    )
    if payload.get("meta", {}).get("error"):
        return JSONResponse(status_code=400, content=payload)
    if billing_user:
        c = cost_research_run()
        if c > 0:
            consume_credits(db, user_id=billing_user.id, amount=c, reason="keyword_research_get", note=seed_keyword[:200])
            db.commit()
    return JSONResponse(content=payload)


@router.post("/research/run")
def api_keyword_research_run(
    body: ResearchRunBody,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    billing_user: User | None = Depends(get_billing_user),
) -> JSONResponse:
    seeds = [str(s).strip() for s in body.seeds if str(s).strip()][:50]
    if not seeds:
        raise HTTPException(status_code=400, detail="At least one keyword required")
    payload = build_keyword_research_api_response(
        seed_keywords=seeds,
        domain=body.domain,
        url=body.url,
        engine=body.engine,
        language=body.language,
        country=body.country,
        device=body.device,
        cluster_mode=body.cluster_mode,
        cluster_strictness=body.cluster_strictness,
        cluster_fetch_serp=body.cluster_fetch_serp,
        cluster_max_keywords=body.cluster_max_keywords,
    )
    if payload.get("meta", {}).get("error"):
        return JSONResponse(status_code=400, content=payload)
    if billing_user:
        c = cost_research_run()
        if c > 0:
            consume_credits(
                db,
                user_id=billing_user.id,
                amount=c,
                reason="keyword_research_run",
                note=f"seeds={len(seeds)}",
            )
    label = " + ".join(seeds[:10])
    if len(seeds) > 10:
        label += " + …"
    row = KeywordResearchProject(
        user_id=current_user.id,
        keywords_label=label[:2000],
        seeds_json=json.dumps(seeds, ensure_ascii=False),
        language=body.language,
        country=body.country,
        engine=body.engine,
        location_label=_country_label(body.country),
        status="completed",
        result_json=json.dumps(payload, ensure_ascii=False),
    )
    db.add(row)
    db.commit()
    db.refresh(row)
    payload["meta"]["project_id"] = row.id
    return JSONResponse(content=payload)


@router.get("/research/history")
def api_research_history(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    limit: int = Query(50, ge=1, le=200),
) -> JSONResponse:
    rows = (
        db.query(KeywordResearchProject)
        .filter(KeywordResearchProject.user_id == current_user.id)
        .order_by(desc(KeywordResearchProject.created_at))
        .limit(limit)
        .all()
    )
    out: list[dict] = []
    for r in rows:
        out.append(
            {
                "id": r.id,
                "keywords_label": r.keywords_label,
                "language": r.language,
                "location_label": r.location_label or _country_label(r.country),
                "status": r.status,
                "created_at": r.created_at.isoformat() if r.created_at else "",
            }
        )
    return JSONResponse(content={"projects": out})


@router.get("/research/history/{project_id}")
def api_research_history_item(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> JSONResponse:
    r = db.get(KeywordResearchProject, project_id)
    if not r or r.user_id != current_user.id or not r.result_json:
        raise HTTPException(status_code=404, detail="Project not found")
    return JSONResponse(content=json.loads(r.result_json))


@router.delete("/research/history/{project_id}")
def api_research_history_delete(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> JSONResponse:
    r = db.get(KeywordResearchProject, project_id)
    if not r or r.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(r)
    db.commit()
    return JSONResponse(content={"ok": True})


@router.post("/research/history/{project_id}/delete")
def api_research_history_delete_post(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> JSONResponse:
    """
    Backward-compatible delete endpoint for clients/proxies that block DELETE.
    """
    r = db.get(KeywordResearchProject, project_id)
    if not r or r.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(r)
    db.commit()
    return JSONResponse(content={"ok": True})


@router.post("/research/delete/{project_id}")
def api_research_delete_alias(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
) -> JSONResponse:
    """
    Extra alias for compatibility with legacy clients.
    """
    r = db.get(KeywordResearchProject, project_id)
    if not r or r.user_id != current_user.id:
        raise HTTPException(status_code=404, detail="Project not found")
    db.delete(r)
    db.commit()
    return JSONResponse(content={"ok": True})


@router.post("/research/run/async")
def api_keyword_research_run_async(
    body: ResearchRunBody,
    db: Session = Depends(get_db),
    billing_user: User | None = Depends(get_billing_user),
) -> JSONResponse:
    """Queue full research + clustering on Celery (``KEYWORD_RESEARCH_ASYNC_CELERY=1`` + worker)."""
    if os.getenv("KEYWORD_RESEARCH_ASYNC_CELERY", "0").lower() not in ("1", "true", "yes"):
        raise HTTPException(
            status_code=503,
            detail="Async keyword research disabled. Set KEYWORD_RESEARCH_ASYNC_CELERY=1 and run a Celery worker.",
        )
    try:
        from app.workers.keyword_tasks import run_keyword_research_task
    except ImportError as exc:
        raise HTTPException(status_code=503, detail=f"Celery worker module unavailable: {exc}") from exc
    seeds = [str(s).strip() for s in body.seeds if str(s).strip()][:50]
    if not seeds:
        raise HTTPException(status_code=400, detail="At least one keyword required")
    if billing_user:
        c = cost_research_run()
        if c > 0:
            consume_credits(db, user_id=billing_user.id, amount=c, reason="keyword_research_async", note=f"seeds={len(seeds)}")
            db.commit()
    task = run_keyword_research_task.delay(body.model_dump())
    return JSONResponse(
        content={
            "task_id": task.id,
            "state": "QUEUED",
            "poll_url": f"/keywords/research/task/{task.id}",
        }
    )


@router.get("/research/task/{task_id}")
def api_keyword_research_task_status(task_id: str) -> JSONResponse:
    """Poll Celery task result for ``/research/run/async``."""
    try:
        from celery.result import AsyncResult

        from app.queue.celery_app import celery_app
    except ImportError as exc:
        raise HTTPException(status_code=503, detail=str(exc)) from exc
    r = AsyncResult(task_id, app=celery_app)
    if r.state == "PENDING":
        return JSONResponse(content={"task_id": task_id, "state": r.state})
    if r.state == "SUCCESS":
        return JSONResponse(content={"task_id": task_id, "state": r.state, "result": r.result})
    return JSONResponse(content={"task_id": task_id, "state": r.state, "info": str(r.info)})


@router.post("/cluster")
async def api_keyword_cluster(
    request: Request,
    db: Session = Depends(get_db),
    billing_user: User | None = Depends(get_billing_user),
) -> JSONResponse:
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON object required")
    kws = body.get("keywords") or body.get("keyword_list")
    if not isinstance(kws, list) or not kws:
        raise HTTPException(status_code=400, detail="keywords array required")
    strings = [str(x).strip() for x in kws if str(x).strip()]
    fetch_serp = body.get("fetch_serp")
    fetch_on = True if fetch_serp is None else bool(fetch_serp)
    brand_hint = str(body.get("url") or body.get("domain") or "").strip() or None
    serp_country = str(body.get("country") or "").strip() or None
    serp_language = str(body.get("language") or "").strip() or None
    serp_device = str(body.get("device") or "").strip() or None
    strict = body.get("cluster_strictness")
    strict_s = str(strict).strip().lower() if strict is not None else None
    payload = build_keyword_cluster_api_response(
        strings,
        fetch_serp=fetch_on,
        brand_host_hint=brand_hint,
        serp_country=serp_country,
        serp_language=serp_language,
        serp_device=serp_device,
        cluster_strictness=strict_s,
    )
    if billing_user:
        c = cost_cluster_sync()
        if c > 0:
            consume_credits(
                db,
                user_id=billing_user.id,
                amount=c,
                reason="keyword_cluster_sync",
                note=f"kw={len(strings)}",
            )
            db.commit()
    return JSONResponse(content=payload)


@router.post("/cluster/async")
async def api_keyword_cluster_async(
    request: Request,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
    billing_user: User | None = Depends(get_billing_user),
) -> JSONResponse:
    """
    Run clustering in background and expose progress % via polling.

    Returns: { job_id, state, progress, poll_url }
    """
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON object required")
    kws = body.get("keywords") or body.get("keyword_list")
    if not isinstance(kws, list) or not kws:
        raise HTTPException(status_code=400, detail="keywords array required")
    strings = [str(x).strip() for x in kws if str(x).strip()]
    fetch_serp = body.get("fetch_serp")
    fetch_on = True if fetch_serp is None else bool(fetch_serp)
    brand_hint = str(body.get("url") or body.get("domain") or "").strip() or None
    serp_country = str(body.get("country") or "").strip() or None
    serp_language = str(body.get("language") or "").strip() or None
    serp_device = str(body.get("device") or "").strip() or None
    strict = body.get("cluster_strictness")
    strict_s = str(strict).strip().lower() if strict is not None else None

    from app.services.job_store import cleanup_expired, create_job, ensure_job_schema, get_job, mark_stale_jobs_failed, mark_stale_queued_failed, update_job

    ensure_job_schema()
    cleanup_expired(ttl_seconds=int(os.getenv("JOB_TTL_SECONDS", "1800")))
    mark_stale_jobs_failed(stale_seconds=int(os.getenv("JOB_WATCHDOG_SECONDS", "300")))
    mark_stale_queued_failed(stale_seconds=int(os.getenv("JOB_QUEUE_STALE_SECONDS", "90")))

    job_payload = {
        "user_id": current_user.id,
        "keywords": strings,
        "fetch_serp": fetch_on,
        "brand_host_hint": brand_hint,
        "serp_country": serp_country,
        "serp_language": serp_language,
        "serp_device": serp_device,
        "cluster_strictness": strict_s,
    }
    job = create_job(job_type="keyword_cluster", message="Queued", payload=job_payload)
    if billing_user:
        c = cost_cluster_async()
        if c > 0:
            consume_credits(
                db,
                user_id=billing_user.id,
                amount=c,
                reason="keyword_cluster_async",
                note=f"kw={len(strings)}",
            )
            db.commit()
    st0 = get_job(job.job_id)
    return JSONResponse(
        content={
            "job_id": job.job_id,
            "state": st0.state if st0 else job.state,
            "progress": st0.progress if st0 else job.progress,
            "poll_url": f"/keywords/cluster/job/{job.job_id}",
        }
    )


@router.get("/cluster/job/{job_id}")
def api_keyword_cluster_job(
    job_id: str,
    current_user: User = Depends(get_current_user),
) -> JSONResponse:
    from app.services.job_store import cleanup_expired, ensure_job_schema, get_job, mark_stale_jobs_failed, mark_stale_queued_failed

    ensure_job_schema()
    cleanup_expired(ttl_seconds=int(os.getenv("JOB_TTL_SECONDS", "1800")))
    mark_stale_jobs_failed(stale_seconds=int(os.getenv("JOB_WATCHDOG_SECONDS", "300")))
    mark_stale_queued_failed(stale_seconds=int(os.getenv("JOB_QUEUE_STALE_SECONDS", "90")))
    st = get_job(job_id)
    if not st:
        raise HTTPException(status_code=404, detail="Job not found")
    assert_job_access(st.payload, current_user.id)
    out = {
        "job_id": st.job_id,
        "state": st.state,
        "progress": st.progress,
        "message": st.message,
    }
    if st.state == "SUCCESS" and st.result is not None:
        out["result"] = st.result
    if st.state == "ERROR":
        out["error"] = st.error or "Unknown error"
    return JSONResponse(content=out)


@router.post("/import/excel")
async def api_import_keywords_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    billing_user: User | None = Depends(get_billing_user),
) -> JSONResponse:
    """
    Upload an Excel (.xlsx) and extract keyword list.

    Heuristics:
    - Detect a 'keyword' column from header (keyword/từ khóa/query/kw)
    - Else use first column
    """
    name = (file.filename or "").lower()
    if not name.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx supported")
    raw = await file.read()
    if not raw or len(raw) < 50:
        raise HTTPException(status_code=400, detail="Empty file")
    if len(raw) > 12_000_000:
        raise HTTPException(status_code=413, detail="File too large (max 12MB)")
    try:
        import openpyxl

        wb = openpyxl.load_workbook(filename=BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"Invalid xlsx: {exc}") from exc

    def norm(x: str) -> str:
        return (x or "").strip().lower()

    header_row_idx: int | None = None
    kw_col_idx: int | None = None  # 1-based
    header_needles = ("keyword", "từ khóa", "tu khoa", "query", "kw")

    for r_idx in range(1, 6):
        row = [ws.cell(row=r_idx, column=c).value for c in range(1, min(30, ws.max_column or 1) + 1)]
        vals = [norm(str(v)) for v in row if v is not None]
        if not vals:
            continue
        for c_idx in range(1, len(row) + 1):
            v = row[c_idx - 1]
            if v is None:
                continue
            nv = norm(str(v))
            if any(n in nv for n in header_needles):
                header_row_idx = r_idx
                kw_col_idx = c_idx
                break
        if kw_col_idx is not None:
            break

    if kw_col_idx is None:
        kw_col_idx = 1
        header_row_idx = 1

    out: list[str] = []
    seen: set[str] = set()
    max_rows = 5000
    start_row = (header_row_idx or 1) + 1

    for r_idx in range(start_row, min(start_row + max_rows, (ws.max_row or start_row) + 1)):
        v = ws.cell(row=r_idx, column=kw_col_idx).value
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        if len(s) > 250:
            s = s[:250].strip()
        low = s.lower()
        if r_idx == start_row and any(n in low for n in header_needles):
            continue
        if low in seen:
            continue
        seen.add(low)
        out.append(s)
        if len(out) >= 2000:
            break

    if billing_user:
        c = cost_import_excel()
        if c > 0:
            consume_credits(db, user_id=billing_user.id, amount=c, reason="import_excel", note=f"rows={len(out)}")
            db.commit()
    return JSONResponse(
        content={
            "keywords": out,
            "meta": {"count": len(out), "truncated": len(out) >= 2000, "column_index": kw_col_idx},
        }
    )


@router.post("/export/excel")
async def api_export_excel(request: Request) -> StreamingResponse:
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON object required")
    keywords = body.get("keywords") if isinstance(body.get("keywords"), list) else []
    clusters = body.get("clusters") if isinstance(body.get("clusters"), list) else []
    if not keywords and not clusters:
        raise HTTPException(status_code=400, detail="Provide keywords and/or clusters")
    data = export_to_excel(keywords=keywords, clusters=clusters)
    name = "keyword_export.xlsx"
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@router.post("/export/cluster-flat-excel")
async def api_export_cluster_flat_excel(request: Request) -> StreamingResponse:
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON object required")
    rows = body.get("rows") if isinstance(body.get("rows"), list) else []
    if not rows:
        raise HTTPException(status_code=400, detail="Provide rows")
    data = export_cluster_flat_to_excel(rows=rows)
    name = "keyword_cluster_flat.xlsx"
    return StreamingResponse(
        iter([data]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{name}"'},
    )


@router.post("/export/sheets")
async def api_export_sheets(request: Request) -> JSONResponse:
    body = await request.json()
    if not isinstance(body, dict):
        raise HTTPException(status_code=400, detail="JSON object required")
    keywords = body.get("keywords") if isinstance(body.get("keywords"), list) else []
    clusters = body.get("clusters") if isinstance(body.get("clusters"), list) else []
    title = str(body.get("title") or "").strip() or None
    url, err = export_to_google_sheets(keywords=keywords, clusters=clusters, title=title)
    if err:
        return JSONResponse(status_code=503, content={"ok": False, "detail": err, "spreadsheet_url": None})
    return JSONResponse(content={"ok": True, "spreadsheet_url": url})


@router.get("", response_class=RedirectResponse)
def keyword_hub_page() -> RedirectResponse:
    return RedirectResponse(url="/keywords/tools/research", status_code=302)


@router.get("/tools/research", response_class=HTMLResponse)
def keyword_research_page(request: Request) -> HTMLResponse:
    return templates.TemplateResponse(
        request=request,
        name="keyword_research.html",
        context={"nav_active": "keyword_research", "kw_tool_active": "research"},
    )


@router.get("/tools/clustering", response_class=RedirectResponse)
def keyword_clustering_page() -> RedirectResponse:
    return RedirectResponse(url="/keywords/tools/research?tab=group", status_code=302)
