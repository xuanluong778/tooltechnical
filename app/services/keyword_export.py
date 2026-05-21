"""
Export keyword research + clusters to Excel (.xlsx) or Google Sheets.
"""

from __future__ import annotations

import io
import os
from datetime import datetime, timezone
from typing import Any

from openpyxl import Workbook


def export_to_excel(
    *,
    keywords: list[dict[str, Any]],
    clusters: list[dict[str, Any]],
) -> bytes:
    """
    Sheet **Keywords**: keyword, intent, avg_volume, difficulty.
    Sheet **Clusters**: cluster_id, main_keyword, intent, keywords (comma-separated).
    """
    wb = Workbook()
    ws_kw = wb.active
    ws_kw.title = "Keywords"
    ws_kw.append(
        [
            "keyword",
            "intent",
            "avg_volume",
            "difficulty",
            "word_count",
            "cpc_avg",
            "cpc_min",
            "cpc_max",
            "competition",
        ]
    )
    for row in keywords or []:
        sv = row.get("search_volume") or {}
        avg = sv.get("avg_monthly")
        if avg is None and isinstance(row.get("search_volume"), (int, float)):
            avg = row.get("search_volume")
        ws_kw.append(
            [
                str(row.get("keyword") or ""),
                str(row.get("intent") or ""),
                avg if avg is not None else "",
                row.get("difficulty") if row.get("difficulty") is not None else "",
                row.get("word_count") if row.get("word_count") is not None else "",
                row.get("cpc_avg") if row.get("cpc_avg") is not None else "",
                row.get("cpc_min") if row.get("cpc_min") is not None else "",
                row.get("cpc_max") if row.get("cpc_max") is not None else "",
                str(row.get("competition") or ""),
            ]
        )

    ws_cl = wb.create_sheet("Clusters")
    ws_cl.append(["cluster_id", "main_keyword", "intent", "keywords"])
    for c in clusters or []:
        kws = c.get("keywords") or []
        if kws and isinstance(kws[0], dict):
            joined = ", ".join(str(x.get("keyword") or "") for x in kws if x.get("keyword"))
        else:
            joined = ", ".join(str(x) for x in kws if x)
        ws_cl.append(
            [
                str(c.get("cluster_id") or ""),
                str(c.get("main_keyword") or ""),
                str(c.get("intent") or ""),
                joined,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_cluster_flat_to_excel(*, rows: list[dict[str, Any]]) -> bytes:
    """
    Flat cluster export (1 row = 1 sub-keyword under a main topic).

    Columns (Vietnamese):
    1) Từ khóa chính
    2) Từ khóa phụ
    3) Phân Loại
    4) Volumg seach
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Clusters"
    ws.append(["Từ khóa chính", "Từ khóa phụ", "Phân Loại", "Volumg seach"])
    for r in rows or []:
        ws.append(
            [
                str(r.get("main_keyword") or ""),
                str(r.get("sub_keyword") or ""),
                str(r.get("classification") or ""),
                r.get("search_volume") if r.get("search_volume") is not None else "",
            ]
        )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_to_google_sheets(
    *,
    keywords: list[dict[str, Any]],
    clusters: list[dict[str, Any]],
    title: str | None = None,
) -> tuple[str | None, str | None]:
    """
    Returns ``(spreadsheet_url, error_message)``. Requires service-account JSON path in env.
    """
    path = (os.getenv("GOOGLE_SHEETS_SA_JSON") or os.getenv("GOOGLE_SHEETS_CREDENTIALS_PATH") or "").strip()
    if not path or not os.path.isfile(path):
        return None, "Set GOOGLE_SHEETS_SA_JSON to a service-account JSON file path (Sheets API enabled)."

    try:
        from google.oauth2 import service_account
        from googleapiclient.discovery import build
    except ImportError as e:
        return None, f"google api client missing: {e}"

    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    creds = service_account.Credentials.from_service_account_file(path, scopes=scopes)
    service = build("sheets", "v4", credentials=creds, cache_discovery=False)

    t = title or f"Keyword export {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M')} UTC"
    spreadsheet_body = {
        "properties": {"title": t[:99]},
        "sheets": [
            {"properties": {"title": "Keywords", "gridProperties": {"rowCount": 5000, "columnCount": 12}}},
            {"properties": {"title": "Clusters", "gridProperties": {"rowCount": 2000, "columnCount": 6}}},
        ],
    }
    created = service.spreadsheets().create(body=spreadsheet_body, fields="spreadsheetId,spreadsheetUrl").execute()
    sid = created.get("spreadsheetId")
    surl = created.get("spreadsheetUrl")

    kw_vals: list[list[Any]] = [
        [
            "keyword",
            "intent",
            "avg_volume",
            "difficulty",
            "word_count",
            "cpc_avg",
            "cpc_min",
            "cpc_max",
            "competition",
        ]
    ]
    for row in keywords or []:
        sv = row.get("search_volume") or {}
        avg = sv.get("avg_monthly")
        if avg is None and isinstance(row.get("search_volume"), (int, float)):
            avg = row.get("search_volume")
        kw_vals.append(
            [
                str(row.get("keyword") or ""),
                str(row.get("intent") or ""),
                avg if avg is not None else "",
                row.get("difficulty") if row.get("difficulty") is not None else "",
                row.get("word_count") if row.get("word_count") is not None else "",
                row.get("cpc_avg") if row.get("cpc_avg") is not None else "",
                row.get("cpc_min") if row.get("cpc_min") is not None else "",
                row.get("cpc_max") if row.get("cpc_max") is not None else "",
                str(row.get("competition") or ""),
            ]
        )

    cl_vals: list[list[Any]] = [["cluster_id", "main_keyword", "intent", "keywords"]]
    for c in clusters or []:
        kws = c.get("keywords") or []
        if kws and isinstance(kws[0], dict):
            joined = ", ".join(str(x.get("keyword") or "") for x in kws if x.get("keyword"))
        else:
            joined = ", ".join(str(x) for x in kws if x)
        cl_vals.append(
            [
                str(c.get("cluster_id") or ""),
                str(c.get("main_keyword") or ""),
                str(c.get("intent") or ""),
                joined,
            ]
        )

    service.spreadsheets().values().update(
        spreadsheetId=sid,
        range="Keywords!A1",
        valueInputOption="USER_ENTERED",
        body={"values": kw_vals},
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=sid,
        range="Clusters!A1",
        valueInputOption="USER_ENTERED",
        body={"values": cl_vals},
    ).execute()

    return surl or f"https://docs.google.com/spreadsheets/d/{sid}/edit", None
