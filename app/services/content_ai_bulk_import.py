"""Parse Excel / text files into Content AI bulk keyword rows."""

from __future__ import annotations

import csv
import re
from io import BytesIO, StringIO
from typing import Any

from app.services.content_ai_bulk_parse import normalize_bulk_outline, _normalize_word_count

_MAX_ROWS = 200

_HEADER_ALIASES: dict[str, tuple[str, ...]] = {
    "keyword": (
        "keyword",
        "từ khóa",
        "tu khoa",
        "primary_keyword",
        "kw",
        "key word",
    ),
    "custom_title": ("title", "tiêu đề", "tieu de", "seo title", "custom_title"),
    "custom_description": (
        "description",
        "meta",
        "meta_description",
        "mô tả",
        "mo ta",
        "meta description",
        "custom_description",
    ),
    "custom_outline": ("outline", "outline_content", "dàn ý", "dan y", "custom_outline"),
    "search_volume": ("volume", "search_volume", "search volume", "vol", "search vol"),
    "content_type": ("content_type", "loại content", "loai content", "type", "content type"),
    "target_word_count": (
        "word_count",
        "số từ",
        "so tu",
        "target_word_count",
        "words",
        "độ dài",
        "do dai",
    ),
    "competitor_url": (
        "competitor_url",
        "url đối thủ",
        "url doi thu",
        "competitor",
        "url",
        "link đối thủ",
    ),
}

_CONTENT_TYPE_MAP: dict[str, str] = {
    "": "",
    "blog": "blog",
    "blog / informational": "blog",
    "informational": "blog",
    "landing": "landing",
    "landing dịch vụ": "landing",
    "landing dich vu": "landing",
    "category": "category",
    "trang danh mục": "category",
    "comparison": "comparison",
    "so sánh": "comparison",
    "review": "comparison",
    "howto": "howto",
    "how-to": "howto",
    "hướng dẫn": "howto",
    "local": "local",
    "địa phương": "local",
}


def _norm_header(value: Any) -> str:
    s = str(value or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _map_header_to_field(header: str) -> str | None:
    h = _norm_header(header)
    if not h:
        return None
    for field, aliases in _HEADER_ALIASES.items():
        if h in aliases:
            return field
        for alias in aliases:
            if alias in h or h in alias:
                return field
    return None


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    return str(value).strip()


def _normalize_content_type(raw: str) -> str:
    key = _norm_header(raw)
    return _CONTENT_TYPE_MAP.get(key, key if key in _CONTENT_TYPE_MAP.values() else "")


def _row_from_mapping(values: dict[str, str]) -> dict[str, Any] | None:
    kw = str(values.get("keyword") or "").strip()
    if not kw:
        return None
    outline_raw = str(values.get("custom_outline") or "").strip()
    try:
        sv = int(float(str(values.get("search_volume") or "0").replace(",", "")))
    except (TypeError, ValueError):
        sv = 0
    return {
        "keyword": re.sub(r"\s+", " ", kw),
        "custom_title": str(values.get("custom_title") or "").strip(),
        "custom_description": str(values.get("custom_description") or "").strip(),
        "custom_outline": normalize_bulk_outline(outline_raw) if outline_raw else "",
        "search_volume": max(0, sv),
        "content_type": _normalize_content_type(str(values.get("content_type") or "")),
        "target_word_count": _normalize_word_count(values.get("target_word_count")),
        "competitor_url": str(values.get("competitor_url") or "").strip(),
    }


def parse_bulk_rows_from_table(rows: list[list[Any]]) -> list[dict[str, Any]]:
    """rows: list of cell lists (first row may be header)."""
    if not rows:
        return []
    header_map: dict[int, str] = {}
    start_idx = 0
    first = rows[0]
    for col_idx, cell in enumerate(first):
        field = _map_header_to_field(_cell_str(cell))
        if field:
            header_map[col_idx] = field
    if header_map:
        start_idx = 1
    else:
        header_map = {0: "keyword", 1: "custom_title", 2: "custom_description", 3: "custom_outline"}
        if len(first) > 4:
            header_map[4] = "search_volume"
        if len(first) > 5:
            header_map[5] = "content_type"
        if len(first) > 6:
            header_map[6] = "target_word_count"
        if len(first) > 7:
            header_map[7] = "competitor_url"

    out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in rows[start_idx:]:
        if len(out) >= _MAX_ROWS:
            break
        values: dict[str, str] = {}
        for col_idx, field in header_map.items():
            if col_idx >= len(row):
                continue
            values[field] = _cell_str(row[col_idx])
        if not values.get("keyword") and row:
            values["keyword"] = _cell_str(row[0])
        parsed = _row_from_mapping(values)
        if not parsed:
            continue
        lk = parsed["keyword"].lower()
        if lk in seen:
            continue
        seen.add(lk)
        out.append(parsed)
    return out


def parse_bulk_text_content(text: str, *, delimiter: str | None = None) -> list[dict[str, Any]]:
    raw = str(text or "").replace("\r\n", "\n").replace("\r", "\n")
    if not raw.strip():
        return []

    lines = [ln for ln in raw.split("\n") if ln.strip()]
    if not lines:
        return []

    first = lines[0]
    if "|" in first and delimiter != ",":
        from app.services.content_ai_bulk_parse import parse_bulk_input_text

        return [
            {
                "keyword": r["keyword"],
                "custom_title": r.get("custom_title") or "",
                "custom_description": r.get("custom_description") or "",
                "custom_outline": r.get("custom_outline") or "",
                "search_volume": 0,
                "content_type": "",
                "target_word_count": 1000,
                "competitor_url": "",
            }
            for r in parse_bulk_input_text(raw)
        ][: _MAX_ROWS]

    sep = delimiter
    if sep is None:
        if "\t" in first:
            sep = "\t"
        elif ";" in first and first.count(";") >= first.count(","):
            sep = ";"
        elif "," in first:
            sep = ","

    if sep:
        reader = csv.reader(StringIO(raw), delimiter=sep)
        table = list(reader)
        return parse_bulk_rows_from_table(table)

    from app.services.content_ai_bulk_parse import parse_bulk_input_text

    return [
        {
            "keyword": r["keyword"],
            "custom_title": r.get("custom_title") or "",
            "custom_description": r.get("custom_description") or "",
            "custom_outline": r.get("custom_outline") or "",
            "search_volume": 0,
            "content_type": "",
            "target_word_count": 1000,
            "competitor_url": "",
        }
        for r in parse_bulk_input_text(raw)
    ][: _MAX_ROWS]


def parse_bulk_xlsx_bytes(raw: bytes) -> list[dict[str, Any]]:
    try:
        import openpyxl
    except ImportError as exc:
        raise ValueError("Thiếu thư viện openpyxl để đọc Excel.") from exc
    if not raw or len(raw) < 50:
        return []
    wb = openpyxl.load_workbook(filename=BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    table: list[list[Any]] = []
    max_r = min((ws.max_row or 0) + 1, _MAX_ROWS + 5)
    max_c = min((ws.max_column or 0) + 1, 20)
    for r_idx in range(1, max_r):
        row = [ws.cell(row=r_idx, column=c).value for c in range(1, max_c)]
        if any(v is not None and str(v).strip() for v in row):
            table.append(row)
    wb.close()
    return parse_bulk_rows_from_table(table)


def parse_bulk_upload_file(*, filename: str, raw: bytes) -> tuple[list[dict[str, Any]], str]:
    name = (filename or "").lower().strip()
    if name.endswith(".xlsx"):
        return parse_bulk_xlsx_bytes(raw), "xlsx"
    if name.endswith(".txt") or name.endswith(".csv"):
        text = raw.decode("utf-8", errors="replace")
        delim = "," if name.endswith(".csv") else None
        return parse_bulk_text_content(text, delimiter=delim), "text"
    if name.endswith(".xls"):
        raise ValueError("File .xls cũ chưa hỗ trợ — vui lòng lưu lại định dạng .xlsx.")
    raise ValueError("Chỉ hỗ trợ .xlsx, .txt hoặc .csv.")
